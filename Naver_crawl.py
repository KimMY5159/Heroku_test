import os
import urllib.request as req
import openpyxl
from bs4 import BeautifulSoup
import time
import re

start = time.time()

# 하트수 em.u_cnt
# 연령  span.age

id_list, address, img_add, latest_add = [], [], [], []
title, author, detail, latest_title = [], [], [], []
genre, rating, is_rest, day = [], [], [], []
now_id = 0
# 연재중 요일별 웹툰 가져오기
week = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
for today in week:
    url = f"https://comic.naver.com/webtoon/weekdayList?week={today}"
    code = req.urlopen(url)
    soup = BeautifulSoup(code, "html.parser")
    link = soup.select("div.list_area.daily_img div.thumb > a")
    rating_tmp = 0
    for i in link:
        j = "https://comic.naver.com" + i.attrs["href"]
        code_each = req.urlopen(j)
        soup_each = BeautifulSoup(code_each, "html.parser")
        t_tmp = soup_each.select_one("h2 > span.title").text
        d_tmp = soup_each.select_one("ul.category_tab > li.on").text[0:1]
        if t_tmp in title:
            day[title.index(t_tmp)] += ', ' + d_tmp
            rating_tmp += 1
            continue
        id_list.append(re.sub(r'[^0-9]', '', j))
        address.append(j)
        title.append(t_tmp)
        day.append(d_tmp)
        img_add.append(soup_each.select_one("div.thumb > a > img").attrs["src"])
        latest_add.append("https://comic.naver.com"+soup_each.select_one("td.title > a").attrs["href"])
        latest_title.append(soup_each.select_one("td.title > a").text)
        author.append(soup_each.select_one("h2 > span.wrt_nm").text)
        detail.append(soup_each.select_one("div.detail > p").text)
        genre.append(soup_each.select_one("span.genre").text)
        rating.append(soup.select("div.list_area.daily_img dl > dd strong")[rating_tmp].text)
        if soup_each.select_one("span.ico_break") is not None:
            is_rest.append("True")
        else:
            is_rest.append("False")
        print(t_tmp)
        now_id += 1
        rating_tmp += 1
    time.sleep(1)

if not os.path.exists("./웹툰크롤링.xlsx"):
    openpyxl.Workbook().save("./웹툰크롤링.xlsx")
book = openpyxl.load_workbook("./웹툰크롤링.xlsx")
# 쓸데없는 시트 삭제하기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
sheet = book.create_sheet()
sheet.title = '네이버_연재중'
# sheet.column_dimensions["A"].width = 15
# sheet.column_dimensions["B"].width = 26
# sheet.column_dimensions["C"].width = 25
# sheet.column_dimensions["D"].width = 60
row_num = 2
for i in range(0, len(title)):
    sheet.cell(row=row_num, column=1).value = id_list[i]
    sheet.cell(row=row_num, column=2).value = title[i]
    sheet.cell(row=row_num, column=3).value = author[i]
    sheet.cell(row=row_num, column=4).value = genre[i]
    sheet.cell(row=row_num, column=5).value = day[i]
    sheet.cell(row=row_num, column=6).value = detail[i]
    sheet.cell(row=row_num, column=7).value = rating[i]
    sheet.cell(row=row_num, column=8).value = is_rest[i]
    sheet.cell(row=row_num, column=9).value = address[i]
    sheet.cell(row=row_num, column=10).value = img_add[i]
    sheet.cell(row=row_num, column=11).value = latest_add[i]
    sheet.cell(row=row_num, column=12).value = latest_title[i]
    row_num += 1
book.save("./웹툰크롤링.xlsx")

print('실행시간 : ', time.time() - start)
